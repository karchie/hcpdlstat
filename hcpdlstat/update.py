# Copyright (c) 2013 Washington University School of Medicine
# Author: Kevin A. Archie <karchie@wustl.edu>

import argparse, datetime, os
import ConfigParser
from openpyxl import load_workbook
import pymysql
import packagelog
import asperastatscollector as aspera

_cfg_reporting = 'reporting'
_cfg_packagelog = 'packagelog'
_cfg_statscollector = 'asperastatscollector'

stats_columns = ['date',
                 'completed_sessions',
                 'completed_users',
                 'completed_bytes',
                 None,
                 'cancelled_sessions',
                 'cancelled_users',
                 None,
                 'error_sessions',
                 'error_users']

pkgs_columns = ['date',
                'g1', 'g1_files',
                'g5', 'g5_files',
                'g20', 'g20_files',
                'g20_avg',
                'preproc', None,
                'unproc', None]

# Some of the packages worksheet values are retrieved from the
# Aspera stats collector database. This is the map of aspera stats
# keys to packages worksheet columns.
_add_pkgs_columns = {
    'preproc_completed':9,
    'unproc_completed':11
}

def cell_from_last_row(worksheet, column):
    """Return the cell from the given column in the last row of the
    given worksheet."""
    return worksheet.cell(row=worksheet.get_highest_row()-1,column=column)

def named_cell_from_last_row(worksheet, names, name):
    """Return the cell from the named column in the last row of the
    given worksheet, given the provided list of column names."""
    return worksheet.cell(row=worksheet.get_highest_row()-1,
                          column=names.index(name))

def get_last_date(s_stats, s_pkgs, file_name):
    """Verifies that the two spreadsheets have matching last dates, and
    return the datetime.date for that last date entry. Raises an Exception
    if the dates do not match."""
    statsdate = named_cell_from_last_row(s_stats, stats_columns, 'date').value.date()
    pkgsdate = named_cell_from_last_row(s_pkgs, pkgs_columns, 'date').value.date()
    if statsdate == pkgsdate:
        return statsdate
    else:
        raise Exception(
"""last date of {} ({}) != last date of {} ({}) in {};
manual intervention is required
""".format(s_stats.title, statsdate, s_pkgs.title, pkgsdate, file_name))
    
def set_row(worksheet, row, values):
    """Sets values in the given row from the provided list."""
    for i, v in enumerate(values):
        worksheet.cell(row=row,column=i).value = v

def set_row_named(worksheet, row, column_names, values):
    """Sets values in the given row from the ordered list of value keys
    and the values dict."""
    set_row(worksheet, row,
            [values[k] if k else k for k in column_names])

def append_row_named(worksheet, keys, valdict):
    """Appends a new row from the ordered list of value keys and the values
    dict. Returns the index of the new row."""
    row = worksheet.get_highest_row()
    set_row_named(worksheet, row, keys, valdict)
    return row

def main():
    argparser = argparse.ArgumentParser(description='Extract statistics from XNAT package request log and Aspera stats collector database into an Excel spreadsheet.')
    argparser.add_argument('file', nargs=1)
    args = argparser.parse_args()
    wb_file_name = args.file[0]

    config = ConfigParser.ConfigParser()
    config.read(['site.cfg', os.path.expanduser('~/.hcpdlstat.cfg')])

    logdir = config.get(_cfg_packagelog, 'logdir')
    logname = config.get(_cfg_packagelog, 'logname')

    stats_name = config.get(_cfg_reporting, 'sheet.stats')
    pkgs_name = config.get(_cfg_reporting, 'sheet.packages')

    wb = load_workbook(wb_file_name)
    s_stats = wb.get_sheet_by_name(stats_name)
    s_pkgs = wb.get_sheet_by_name(pkgs_name)

    host = config.get(_cfg_statscollector, 'mysql.host')
    port = int(config.get(_cfg_statscollector, 'mysql.port'))
    user = config.get(_cfg_statscollector, 'mysql.user')
    passwd = config.get(_cfg_statscollector, 'mysql.password')
    db = pymysql.connect(host=host,port=port,user=user,passwd=passwd)

    date = get_last_date(s_stats, s_pkgs, wb_file_name)

    # add rows up to (but excluding) today
    while True:
        date += datetime.timedelta(days=1)
        if date >= datetime.date.today():
            break

        pkgstats = packagelog.get_stats(logdir, logname, date)
        pkgrow = append_row_named(s_pkgs, pkgs_columns, pkgstats)

        filestats = aspera.get_stats(db, date)
        filesrow = append_row_named(s_stats, stats_columns, filestats)
        
        # Some of thep
        for k,col in _add_pkgs_columns.iteritems():
            s_pkgs.cell(row=pkgrow,column=col).value=filestats[k]

    # write the modified spreadsheet file
    wb.save(wb_file_name)
